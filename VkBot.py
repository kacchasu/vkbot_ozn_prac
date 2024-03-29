import requests
import bs4 as bs
import vk_api
from vk_api import VkUpload
from vk_api.keyboard import VkKeyboard, VkKeyboardColor
from vk_api.utils import get_random_id
import openpyxl
import re
import datetime
from vk_api.longpoll import VkLongPoll, VkEventType
from Schedule import Schedule
import shelve
from Corona import Corona
from WeatherProvider import WeatherProvider, WeatherObject
import PIL.Image as Image


VK_API_TOKEN = '549032f8f34c3497616a86b50173efb07243d0910e25de0f5ae1212072292f46a76851ba7122662ffdd78'

class VkBot:
    def __init__(self, user_id):
        self.logging("CRE", "bot class object created")
        get_schedule_files()

        self._USER_ID = user_id

        with shelve.open("groups.txt") as c_groups:
            if str(self._USER_ID) in c_groups:
                self.group = c_groups[str(self._USER_ID)]
            else:
                self.group = "не установлена"

        self.day = datetime.datetime.today().weekday()
        self._USERNAME = self._get_user_name_from_vk_id(user_id)

    # for logging
    @staticmethod
    def logging(key, comm):
        f1 = open("log.txt", "a", encoding='UTF-8')
        t1 = datetime.datetime.now()
        f1.write(key + " --- ")
        f1.write(t1.strftime("%d/%m/%Y %H:%M") + " --- ")
        f1.write("vkbot class: " + comm + '\n')

    # for getting username
    def _get_user_name_from_vk_id(self, user_id):
        request = requests.get("https://vk.com/id" + str(user_id))
        soup = bs.BeautifulSoup(request.text, "html.parser")

        user_name = self._clean_all_tag_from_str(soup.findAll("title")[0])

        return user_name.split()[0]

    # new message handler
    def new_message(self, message):

        # start
        if message.upper() == "НАЧАТЬ" or message.upper() == "ПРИВЕТ" or message.upper() == "/HELP":
            self.logging("SENT", "start message")
            vk.messages.send(
                user_id=event.user_id,
                random_id=get_random_id(),
                message=f"добро пожаловать {self._USERNAME.lower()}!\n" + "список команд и описание работы бота:\n" \
                        + "---\n" \
                        + "<ИМЯ_ГРУППЫ> - запомнить заданную группу как текущую\n" \
                        + "---\n" \
                        + "бот - показать расписание для текущей группы\n" \
                        + "---\n" \
                        + "бот <ИМЯ_ГРУППЫ> - показать расписание для <ИМЯ_ГРУППЫ>\n" \
                        + "---\n" \
                        + "бот <ДЕНЬ_НЕДЕЛИ> - показать расписание текущей группы на четный\нечетный <ДЕНЬ_НЕДЕЛИ>\n" \
                        + "---\n" \
                        + "бот <ДЕНЬ_НЕДЕЛИ> <ИМЯ_ГРУППЫ> - показать расписание <ИМЯ_ГРУППЫ> на четный\нечетный <ДЕНЬ_НЕДЕЛИ>\n" \
                        + "---\n" \
                        + "погода - показать погоду в москве: состояние погоды, температура, давление в мм рт. ст., влажность, описание, сила и направление ветра\n" \
                        + "---\n" \
                        + "найти <ФАМИЛИЯ_ПРЕПОДАВАТЕЛЯ> - показать расписание преподавателя\n" \
                        + "---\n" \
                        + "корона - показать статистику на текущий день и график за предыдущие 10 дней\n" \
                        + "---\n" \
                        + "корона <РЕГИОН> - статистика на текущий день в <РЕГИОН>\n" \
                        + "---\n" \
                        + "/help для повтора данного сообщения"
            )

        # group tag
        elif re.fullmatch(r"[А-Я][А-Я][А-Я][А-Я][-]\d\d[-]\d\d", message.upper()):
            with shelve.open("groups.txt") as groups:
                groups[str(self._USER_ID)] = message.upper()
            self.logging("UPD", "current group updated: " + message)
            vk.messages.send(
                user_id=event.user_id,
                random_id=get_random_id(),
                message="текущая группа установлена: " + message
            )

        # get schedule for current group
        elif message.upper() == "БОТ":
            with shelve.open("groups.txt") as c_groups:
                if str(self._USER_ID) in c_groups:
                    self.set_keyboard_schedule(1, c_groups[str(self._USER_ID)])

                    schedule = Schedule(c_groups[str(self._USER_ID)], "", course1, course2, course3)

                    longp = VkLongPoll(vk_session)

                    for event_c in longpoll.listen():
                        if event_c.type == VkEventType.MESSAGE_NEW:
                            if event_c.to_me:
                                sch = schedule.get_week_schedule()
                                message = ''
                                if event_c.text == "на сегодня":
                                    if self.day == 6:
                                        vk.messages.send(
                                            user_id=event.user_id,
                                            random_id=get_random_id(),
                                            message="сегодня воскресенье"
                                        )
                                        break
                                    message += f"расписание {c_groups[str(self._USER_ID)]} на {['понедельник', 'вторник', 'среду', 'четверг', 'пятницу', 'субботу'][self.day]}\n"
                                    if schedule.week % 2 == 1:
                                        j = 1
                                        for i in range(0, 12, 2):
                                            message += f"{j}) {sch[i + self.day * 12]}\n"
                                            j += 1
                                    else:
                                        j = 1
                                        for i in range(1, 12, 2):
                                            message += f"{j}) {sch[i + self.day * 12]}\n"
                                            j += 1
                                    vk.messages.send(
                                        user_id=event.user_id,
                                        random_id=get_random_id(),
                                        message=message
                                    )
                                    break
                                elif event_c.text == "на завтра":
                                    self.day = (self.day + 1) % 7
                                    if self.day == 6:
                                        vk.messages.send(
                                            user_id=event.user_id,
                                            random_id=get_random_id(),
                                            message="завтра воскресенье"
                                        )
                                        break
                                    message += f"расписание {c_groups[str(self._USER_ID)]} на {['понедельник', 'вторник', 'среду', 'четверг', 'пятницу', 'субботу'][self.day]}\n"
                                    if self.day == 0:
                                        schedule.week += 1
                                    if schedule.week % 2 == 1:
                                        j = 1
                                        for i in range(0, 12, 2):
                                            message += f"{j}) {sch[i + self.day * 12]}\n"
                                            j += 1
                                    else:
                                        j = 1
                                        for i in range(1, 12, 2):
                                            message += f"{j}) {sch[i + self.day * 12]}\n"
                                            j += 1
                                    vk.messages.send(
                                        user_id=event.user_id,
                                        random_id=get_random_id(),
                                        message=message
                                    )
                                    break
                                elif event_c.text == "на эту неделю":
                                    message += f"расписание {c_groups[str(self._USER_ID)]} на {schedule.week} неделю\n"
                                    if schedule.week % 2 == 1:
                                        for d in range(6):
                                            j = 1
                                            message += \
                                                ["понедельник", 'вторник', "среда", "четверг", "пятница", "суббота"][
                                                    d] + "\n"
                                            for i in range(0, 12, 2):
                                                message += f"{j}) {sch[i + d * 12]}\n"
                                                j += 1
                                    else:
                                        for d in range(6):
                                            j = 1
                                            message += \
                                                ["понедельник", 'вторник', "среда", "четверг", "пятница", "суббота"][
                                                    d] + "\n"
                                            for i in range(1, 12, 2):
                                                message += f"{j}) {sch[i + d * 12]}\n"
                                                j += 1
                                    vk.messages.send(
                                        user_id=event.user_id,
                                        random_id=get_random_id(),
                                        message=message
                                    )
                                    break
                                elif event_c.text == "на следующую неделю":
                                    schedule.week += 1
                                    message += f"расписание {c_groups[str(self._USER_ID)]} на {schedule.week} неделю\n"
                                    if schedule.week % 2 == 1:
                                        for d in range(6):
                                            j = 1
                                            message += \
                                                ["понедельник", 'вторник', "среда", "четверг", "пятница", "суббота"][
                                                    d] + "\n"
                                            for i in range(0, 12, 2):
                                                message += f"{j}) {sch[i + d * 12]}\n"
                                                j += 1
                                    else:
                                        for d in range(6):
                                            j = 1
                                            message += \
                                                ["понедельник", 'вторник', "среда", "четверг", "пятница", "суббота"][
                                                    d] + "\n"
                                            for i in range(1, 12, 2):
                                                message += f"{j}) {sch[i + d * 12]}\n"
                                                j += 1
                                    vk.messages.send(
                                        user_id=event.user_id,
                                        random_id=get_random_id(),
                                        message=message
                                    )
                                    break
                                elif event_c.text == "какая группа?":
                                    vk.messages.send(
                                        user_id=event.user_id,
                                        random_id=get_random_id(),
                                        message=f"текущая группа: {c_groups[str(self._USER_ID)]}"
                                    )
                                    break
                                elif event_c.text == "какая неделя?":
                                    vk.messages.send(
                                        user_id=event.user_id,
                                        random_id=get_random_id(),
                                        message=f"идет {schedule.week} неделя"
                                    )
                                    break
                                else:
                                    vk.messages.send(
                                        user_id=event.user_id,
                                        random_id=get_random_id(),
                                        message="некорректный запрос"
                                    )
                                    break
                else:
                    vk.messages.send(
                        user_id=event.user_id,
                        random_id=get_random_id(),
                        message="текущая группа не задана"
                    )
            self.logging("SENT", "БОТ msg handled")

        # get schedule for any group
        elif re.fullmatch("БОТ " + r"[А-Я][А-Я][А-Я][А-Я][-]\d\d[-]\d\d", message.upper()):
            self.set_keyboard_schedule(1, message[4:len(message)].upper())

            schedule = Schedule(message[4:len(message)].upper(), "", course1, course2, course3)

            longp = VkLongPoll(vk_session)

            for event_c in longpoll.listen():
                if event_c.type == VkEventType.MESSAGE_NEW:
                    if event_c.to_me:
                        sch = schedule.get_week_schedule()
                        message = ''
                        if event_c.text == "на сегодня":
                            if self.day == 6:
                                vk.messages.send(
                                    user_id=event.user_id,
                                    random_id=get_random_id(),
                                    message="сегодня воскресенье"
                                )
                                break
                            message += f"расписание {schedule.group} на {['понедельник', 'вторник', 'среду', 'четверг', 'пятницу', 'субботу'][self.day]}\n"
                            if schedule.week % 2 == 1:
                                j = 1
                                for i in range(0, 12, 2):
                                    message += f"{j}) {sch[i + self.day * 12]}\n"
                                    j += 1
                            else:
                                j = 1
                                for i in range(1, 12, 2):
                                    message += f"{j}) {sch[i + self.day * 12]}\n"
                                    j += 1
                            vk.messages.send(
                                user_id=event.user_id,
                                random_id=get_random_id(),
                                message=message
                            )
                            break
                        elif event_c.text == "на завтра":
                            self.day = (self.day + 1) % 7
                            if self.day == 6:
                                vk.messages.send(
                                    user_id=event.user_id,
                                    random_id=get_random_id(),
                                    message="завтра воскресенье"
                                )
                                break
                            message += f"расписание {schedule.group} на {['понедельник', 'вторник', 'среду', 'четверг', 'пятницу', 'субботу'][self.day]}\n"
                            if self.day == 0:
                                schedule.week += 1
                            if schedule.week % 2 == 1:
                                j = 1
                                for i in range(0, 12, 2):
                                    message += f"{j}) {sch[i + self.day * 12]}\n"
                                    j += 1
                            else:
                                j = 1
                                for i in range(1, 12, 2):
                                    message += f"{j}) {sch[i + self.day * 12]}\n"
                                    j += 1
                            vk.messages.send(
                                user_id=event.user_id,
                                random_id=get_random_id(),
                                message=message
                            )
                            break
                        elif event_c.text == "на эту неделю":
                            message += f"расписание {schedule.group} на {schedule.week} неделю\n"
                            if schedule.week % 2 == 1:
                                for d in range(6):
                                    j = 1
                                    message += \
                                        ["понедельник", 'вторник', "среда", "четверг", "пятница", "суббота"][
                                            d] + "\n"
                                    for i in range(0, 12, 2):
                                        message += f"{j}) {sch[i + d * 12]}\n"
                                        j += 1
                            else:
                                for d in range(6):
                                    j = 1
                                    message += \
                                        ["понедельник", 'вторник', "среда", "четверг", "пятница", "суббота"][
                                            d] + "\n"
                                    for i in range(1, 12, 2):
                                        message += f"{j}) {sch[i + d * 12]}\n"
                                        j += 1
                            vk.messages.send(
                                user_id=event.user_id,
                                random_id=get_random_id(),
                                message=message
                            )
                            break
                        elif event_c.text == "на следующую неделю":
                            schedule.week += 1
                            message += f"расписание {schedule.group} на {schedule.week} неделю\n"
                            if schedule.week % 2 == 1:
                                for d in range(6):
                                    j = 1
                                    message += \
                                        ["понедельник", 'вторник', "среда", "четверг", "пятница", "суббота"][
                                            d] + "\n"
                                    for i in range(0, 12, 2):
                                        message += f"{j}) {sch[i + d * 12]}\n"
                                        j += 1
                            else:
                                for d in range(6):
                                    j = 1
                                    message += \
                                        ["понедельник", 'вторник', "среда", "четверг", "пятница", "суббота"][
                                            d] + "\n"
                                    for i in range(1, 12, 2):
                                        message += f"{j}) {sch[i + d * 12]}\n"
                                        j += 1
                            vk.messages.send(
                                user_id=event.user_id,
                                random_id=get_random_id(),
                                message=message
                            )
                            break
                        elif event_c.text == "какая группа?":
                            vk.messages.send(
                                user_id=event.user_id,
                                random_id=get_random_id(),
                                message=f"текущая группа: {self.group}"
                            )
                            break
                        elif event_c.text == "какая неделя?":
                            vk.messages.send(
                                user_id=event.user_id,
                                random_id=get_random_id(),
                                message=f"идет {schedule.week} неделя"
                            )
                            break
                        else:
                            vk.messages.send(
                                user_id=event.user_id,
                                random_id=get_random_id(),
                                message="некорректный запрос"
                            )
                            break
            self.logging("SENT", "БОТ <ГРУППА> msg handled")

        # get schedule for current group on any even\uneven day
        elif re.fullmatch("БОТ " + r"(ПОНЕДЕЛЬНИК|ВТОРНИК|СРЕДА|ЧЕТВЕРГ|ПЯТНИЦА|СУББОТА|ВОСКРЕСЕНЬЕ)", message.upper()):
            with shelve.open("groups.txt") as c_groups:
                if str(self._USER_ID) in c_groups:
                    schedule = Schedule(c_groups[str(self._USER_ID)], "",
                                        course1, course2, course3)
                    sch = schedule.get_week_schedule()
                    longp = VkLongPoll(vk_session)

                    self.day = ['ПОНЕДЕЛЬНИК', 'ВТОРНИК', 'СРЕДА', 'ЧЕТВЕРГ', 'ПЯТНИЦА', 'СУББОТА',
                                'ВОСКРЕСЕНЬЕ'].index(
                        re.findall(r"(ПОНЕДЕЛЬНИК|ВТОРНИК|СРЕДА|ЧЕТВЕРГ|ПЯТНИЦА|СУББОТА|ВОСКРЕСЕНЬЕ)", message.upper())[
                            0])

                    message = ""
                    if self.day == 6:
                        vk.messages.send(
                            user_id=event.user_id,
                            random_id=get_random_id(),
                            message="воскресенье без пар"
                        )
                        return
                    message += f"расписание {schedule.group} на нечет {['понедельник', 'вторник', 'среду', 'четверг', 'пятницу', 'субботу'][self.day]}\n"
                    j = 1
                    for i in range(0, 12, 2):
                        message += f"{j}) {sch[i + self.day * 12]}\n"
                        j += 1
                    message += f"расписание {schedule.group} на чет {['понедельник', 'вторник', 'среду', 'четверг', 'пятницу', 'субботу'][self.day]}\n"
                    j = 1
                    for i in range(1, 12, 2):
                        message += f"{j}) {sch[i + self.day * 12]}\n"
                        j += 1
                    vk.messages.send(
                        user_id=event.user_id,
                        random_id=get_random_id(),
                        message=message
                    )

                else:
                    return "текущая группа не задана"

        # get schedule for any group on any even\uneven day
        elif re.fullmatch(
                r"БОТ (ПОНЕДЕЛЬНИК|ВТОРНИК|СРЕДА|ЧЕТВЕРГ|ПЯТНИЦА|СУББОТА|ВОСКРЕСЕНЬЕ) [А-Я][А-Я][А-Я][А-Я][-]\d\d[-]\d\d",
                message.upper()):

            schedule = Schedule(re.findall(r"[А-Я][А-Я][А-Я][А-Я][-]\d\d[-]\d\d", message.upper())[0], "", course1,
                                course2, course3)
            sch = schedule.get_week_schedule()
            longp = VkLongPoll(vk_session)

            self.day = ['ПОНЕДЕЛЬНИК', 'ВТОРНИК', 'СРЕДА', 'ЧЕТВЕРГ', 'ПЯТНИЦА', 'СУББОТА', 'ВОСКРЕСЕНЬЕ'].index(
                re.findall(r"(ПОНЕДЕЛЬНИК|ВТОРНИК|СРЕДА|ЧЕТВЕРГ|ПЯТНИЦА|СУББОТА|ВОСКРЕСЕНЬЕ)", message.upper())[0])

            message = ""
            if self.day == 6:
                vk.messages.send(
                    user_id=event.user_id,
                    random_id=get_random_id(),
                    message="воскресенье без пар"
                )
                return
            message += f"расписание {schedule.group} на нечет {['понедельник', 'вторник', 'среду', 'четверг', 'пятницу', 'субботу'][self.day]}\n"
            j = 1
            for i in range(0, 12, 2):
                message += f"{j}) {sch[i + self.day * 12]}\n"
                j += 1
            message += f"расписание {schedule.group} на чет {['понедельник', 'вторник', 'среду', 'четверг', 'пятницу', 'субботу'][self.day]}\n"
            j = 1
            for i in range(1, 12, 2):
                message += f"{j}) {sch[i + self.day * 12]}\n"
                j += 1
            vk.messages.send(
                user_id=event.user_id,
                random_id=get_random_id(),
                message=message
            )

        # get weather
        elif message.upper() == "ПОГОДА":
            self.set_keyboard_weather()
            weatherProvider = WeatherProvider()
            upload = VkUpload(vk_session)
            attachments = list()

            longp = VkLongPoll(vk_session)

            for event_c in longpoll.listen():
                if event_c.type == VkEventType.MESSAGE_NEW:
                    if event_c.to_me:
                        message = ''
                        if event_c.text == "сейчас":
                            weatherObject = weatherProvider.get_current_weather()
                            photo = upload.photo_messages(photos = "icons/" + weatherObject.icon + ".png")[0]
                            attachments.append("photo{}_{}".format(photo["owner_id"], photo["id"]))

                            vk.messages.send(
                                user_id=event.user_id,
                                attachment=','.join(attachments),
                                random_id=get_random_id(),
                                message="погода в москве"
                            )

                            vk.messages.send(
                                user_id=event.user_id,
                                random_id=get_random_id(),
                                message=f"{weatherObject.description}, температура: {weatherObject.min_temp}-{weatherObject.max_temp}℃\n\
                                            Давление: {weatherObject.pressure} мм рт. ст., влажность: {weatherObject.humidity}%\n\
                                            Ветер: {weatherObject.wind_type}, {weatherObject.wind_speed} м/с, {weatherObject.direction}"
                            )
                            self.logging("SENT", "sent current weather message")
                            break
                        elif event_c.text == "на сегодня":
                            today_weather_dict = weatherProvider.get_today_weather()
                            morning = today_weather_dict.get("morning", None)
                            day = today_weather_dict.get("day", None)
                            evening = today_weather_dict.get("evening", None)
                            night = today_weather_dict.get("night", None)

                            image_to_send = Image.new("RGBA", (400, 100))
                            current_x_position = 0
                            string_to_send = "/"

                            if morning is not None:
                                icon_to_insert = Image.open("icons/" + morning.icon + ".png")
                                image_to_send.paste(icon_to_insert, (current_x_position, 0))
                                icon_to_insert.close()
                                current_x_position += 100
                                string_to_send += "\t" + str((morning.max_temp + morning.min_temp) // 2) + "℃ //"

                            if day is not None:
                                icon_to_insert = Image.open("icons/" + day.icon + ".png")
                                image_to_send.paste(icon_to_insert, (current_x_position, 0))
                                icon_to_insert.close()
                                current_x_position += 100
                                string_to_send += "\t" + str((day.max_temp + day.min_temp) // 2) + "℃ //"

                            if evening is not None:
                                icon_to_insert = Image.open("icons/" + evening.icon + ".png")
                                image_to_send.paste(icon_to_insert, (current_x_position, 0))
                                icon_to_insert.close()
                                current_x_position += 100
                                string_to_send += "\t" + str((evening.max_temp + evening.min_temp) // 2) + "℃ //"

                            if night is not None:
                                icon_to_insert = Image.open("icons/" + night.icon + ".png")
                                image_to_send.paste(icon_to_insert, (current_x_position, 0))
                                icon_to_insert.close()
                                current_x_position += 100
                                string_to_send += "\t" + str((night.max_temp + night.min_temp) // 2) + "℃ /"

                            image_to_send.save("tmp_icon.png")
                            photo = upload.photo_messages(photos="tmp_icon.png")[0]
                            attachments.append("photo{}_{}".format(photo["owner_id"], photo["id"]))

                            vk.messages.send(
                                user_id=event.user_id,
                                attachment=','.join(attachments),
                                random_id=get_random_id(),
                                message="погода в москве на сегодня"
                            )
                            vk.messages.send(
                                user_id=event.user_id,
                                random_id=get_random_id(),
                                message=string_to_send
                            )

                            if morning is not None:
                                vk.messages.send(
                                    user_id=event.user_id,
                                    random_id=get_random_id(),
                                    message=f"УТРО\n\
                                            // {morning.description}, температура: {morning.min_temp}-{morning.max_temp}℃\n\
                                            // Давление: {morning.pressure} мм рт. ст., влажность: {morning.humidity}%\n\
                                            // Ветер: {morning.wind_type}, {morning.wind_speed} м/с, {morning.direction}"
                                )

                            if day is not None:
                                vk.messages.send(
                                    user_id=event.user_id,
                                    random_id=get_random_id(),
                                    message=f"ДЕНЬ\n\
                                            // {day.description}, температура: {day.min_temp}-{day.max_temp}℃\n\
                                            // Давление: {day.pressure} мм рт. ст., влажность: {day.humidity}%\n\
                                            // Ветер: {day.wind_type}, {day.wind_speed} м/с, {day.direction}"
                                )

                            if evening is not None:
                                vk.messages.send(
                                    user_id=event.user_id,
                                    random_id=get_random_id(),
                                    message=f"ВЕЧЕР\n\
                                            // {evening.description}, температура: {evening.min_temp}-{evening.max_temp}℃\n\
                                            // Давление: {evening.pressure} мм рт. ст., влажность: {evening.humidity}%\n\
                                            // Ветер: {evening.wind_type}, {evening.wind_speed} м/с, {evening.direction}"
                                )

                            if night is not None:
                                vk.messages.send(
                                    user_id=event.user_id,
                                    random_id=get_random_id(),
                                    message=f"НОЧЬ\n\
                                            // {night.description}, температура: {night.min_temp}-{night.max_temp}℃\n\
                                            // Давление: {night.pressure} мм рт. ст., влажность: {night.humidity}%\n\
                                            // Ветер: {night.wind_type}, {night.wind_speed} м/с, {night.direction}"
                                )
                            self.logging("SENT", "sent today weather message")
                            break
                        elif event_c.text == "на завтра":
                            tomorrow_weather_dict = weatherProvider.get_tomorrow_weather()
                            morning = tomorrow_weather_dict.get("morning", None)
                            day = tomorrow_weather_dict.get("day", None)
                            evening = tomorrow_weather_dict.get("evening", None)
                            night = tomorrow_weather_dict.get("night", None)

                            image_to_send = Image.new("RGBA", (400, 100))
                            current_x_position = 0
                            string_to_send = "/"

                            if morning is not None:
                                icon_to_insert = Image.open("icons/" + morning.icon + ".png")
                                image_to_send.paste(icon_to_insert, (current_x_position, 0))
                                icon_to_insert.close()
                                current_x_position += 100
                                string_to_send += "\t" + str((morning.max_temp + morning.min_temp) // 2) + "℃ //"

                            if day is not None:
                                icon_to_insert = Image.open("icons/" + day.icon + ".png")
                                image_to_send.paste(icon_to_insert, (current_x_position, 0))
                                icon_to_insert.close()
                                current_x_position += 100
                                string_to_send += "\t" + str((day.max_temp + day.min_temp) // 2) + "℃ //"

                            if evening is not None:
                                icon_to_insert = Image.open("icons/" + evening.icon + ".png")
                                image_to_send.paste(icon_to_insert, (current_x_position, 0))
                                icon_to_insert.close()
                                current_x_position += 100
                                string_to_send += "\t" + str((evening.max_temp + evening.min_temp) // 2) + "℃ //"

                            if night is not None:
                                icon_to_insert = Image.open("icons/" + night.icon + ".png")
                                image_to_send.paste(icon_to_insert, (current_x_position, 0))
                                icon_to_insert.close()
                                current_x_position += 100
                                string_to_send += "\t" + str((night.max_temp + night.min_temp) // 2) + "℃ /"

                            image_to_send.save("tmp_icon.png")
                            photo = upload.photo_messages(photos="tmp_icon.png")[0]
                            attachments.append("photo{}_{}".format(photo["owner_id"], photo["id"]))

                            vk.messages.send(
                                user_id=event.user_id,
                                attachment=','.join(attachments),
                                random_id=get_random_id(),
                                message="погода в москве на завтра"
                            )
                            vk.messages.send(
                                user_id=event.user_id,
                                random_id=get_random_id(),
                                message=string_to_send
                            )

                            if morning is not None:
                                vk.messages.send(
                                    user_id=event.user_id,
                                    random_id=get_random_id(),
                                    message=f"УТРО\n\
                                            // {morning.description}, температура: {morning.min_temp}-{morning.max_temp}℃\n\
                                            // Давление: {morning.pressure} мм рт. ст., влажность: {morning.humidity}%\n\
                                            // Ветер: {morning.wind_type}, {morning.wind_speed} м/с, {morning.direction}"
                                )

                            if day is not None:
                                vk.messages.send(
                                    user_id=event.user_id,
                                    random_id=get_random_id(),
                                    message=f"ДЕНЬ\n\
                                            // {day.description}, температура: {day.min_temp}-{day.max_temp}℃\n\
                                            // Давление: {day.pressure} мм рт. ст., влажность: {day.humidity}%\n\
                                            // Ветер: {day.wind_type}, {day.wind_speed} м/с, {day.direction}"
                                )

                            if evening is not None:
                                vk.messages.send(
                                    user_id=event.user_id,
                                    random_id=get_random_id(),
                                    message=f"ВЕЧЕР\n\
                                            // {evening.description}, температура: {evening.min_temp}-{evening.max_temp}℃\n\
                                            // Давление: {evening.pressure} мм рт. ст., влажность: {evening.humidity}%\n\
                                            // Ветер: {evening.wind_type}, {evening.wind_speed} м/с, {evening.direction}"
                                )

                            if night is not None:
                                vk.messages.send(
                                    user_id=event.user_id,
                                    random_id=get_random_id(),
                                    message=f"НОЧЬ\n\
                                            // {night.description}, температура: {night.min_temp}-{night.max_temp}℃\n\
                                            // Давление: {night.pressure} мм рт. ст., влажность: {night.humidity}%\n\
                                            // Ветер: {night.wind_type}, {night.wind_speed} м/с, {night.direction}"
                                )
                            self.logging("SENT", "sent tomorrow weather message")
                            break
                        elif event_c.text == "на 5 дней":
                            all_preiod_weather_dict = weatherProvider.get_all_period_weather()
                            image_to_send = Image.new("RGBA", (500, 100))
                            current_x_position = 0
                            string_day_to_send = ""
                            string_night_to_send = ""
                            string_period_to_send = "погода в москве с "

                            current_period_number = 0
                            for key in sorted(all_preiod_weather_dict.keys()):
                                if current_period_number == 0:
                                    string_period_to_send += key.strftime("%d.%m") + " по "
                                elif current_period_number == 4:
                                    string_period_to_send += key.strftime("%d.%m")
                                elif current_period_number > 4:
                                    break

                                current_day_weather_dict = all_preiod_weather_dict[key]
                                current_weather_day = current_day_weather_dict.get("day", None)
                                current_weather_night = current_day_weather_dict.get("night", None)

                                if current_weather_day:
                                    string_day_to_send += f"/ {(current_weather_day.min_temp + current_weather_day.max_temp) // 2}℃ /"
                                else:
                                    string_day_to_send += "/ -- /"

                                if current_weather_night:
                                    string_night_to_send += f"/ {(current_weather_night.min_temp + current_weather_night.max_temp) // 2}℃ /"
                                else:
                                    string_night_to_send += "/ -- /"

                                if current_weather_day:
                                    icon_to_insert = Image.open("icons/" + current_weather_day.icon + ".png")
                                    image_to_send.paste(icon_to_insert, (current_period_number * 100, 0))
                                    icon_to_insert.close()
                                else: # if day's icon is not defined we'll use evening one
                                    icon_to_insert = Image.open("icons/" + current_day_weather_dict["evening"].icon + ".png")
                                    image_to_send.paste(icon_to_insert, (current_period_number * 100, 0))
                                    icon_to_insert.close()

                                current_period_number += 1

                            image_to_send.save("tmp_icon.png")
                            photo = upload.photo_messages(photos="tmp_icon.png")[0]
                            attachments.append("photo{}_{}".format(photo["owner_id"], photo["id"]))

                            vk.messages.send(
                                user_id=event.user_id,
                                attachment=','.join(attachments),
                                random_id=get_random_id(),
                                message=string_period_to_send
                            )
                            vk.messages.send(
                                user_id=event.user_id,
                                random_id=get_random_id(),
                                message=f"{string_day_to_send} ДЕНЬ\n\
                                        {string_night_to_send} НОЧЬ"
                            )
                            self.logging("SENT", "sent all period weather message")
                            break
        # get teacher's schedule
        elif re.fullmatch("НАЙТИ " + r".+", message.upper()):
            schedule = Schedule("", message[6:].title(), course1, course2, course3)
            if schedule.teacher[-1] == ".":
                tchrs = [schedule.teacher]
            else:
                tchrs = list(schedule.check_surnames())
            if len(tchrs) > 1:
                self.set_keyboard_surnames(tchrs)
            elif len(tchrs) == 1:
                schedule.teacher = tchrs[0]
                self.set_keyboard_schedule(0, "", schedule.teacher)

                longp = VkLongPoll(vk_session)

                for event_c in longpoll.listen():
                    if event_c.type == VkEventType.MESSAGE_NEW:

                        if event_c.to_me:
                            sch = schedule.get_week_schedule()
                            message = ''
                            if event_c.text == "на сегодня":
                                if self.day == 6:
                                    vk.messages.send(
                                        user_id=event.user_id,
                                        random_id=get_random_id(),
                                        message="сегодня воскресенье"
                                    )
                                    break
                                message += f"расписание {schedule.teacher} на {['понедельник', 'вторник', 'среду', 'четверг', 'пятницу', 'субботу'][self.day]}\n"
                                if schedule.week % 2 == 1:
                                    j = 1
                                    for i in range(0, 12, 2):
                                        message += f"{j}) {sch[i + self.day * 12]}\n"
                                        j += 1
                                else:
                                    j = 1
                                    for i in range(1, 12, 2):
                                        message += f"{j}) {sch[i + self.day * 12]}\n"
                                        j += 1
                                vk.messages.send(
                                    user_id=event.user_id,
                                    random_id=get_random_id(),
                                    message=message
                                )
                                break
                            elif event_c.text == "на завтра":
                                self.day = (self.day + 1) % 7
                                if self.day == 6:
                                    vk.messages.send(
                                        user_id=event.user_id,
                                        random_id=get_random_id(),
                                        message="завтра воскресенье"
                                    )
                                    break
                                message += f"расписание {schedule.teacher} на {['понедельник', 'вторник', 'среду', 'четверг', 'пятницу', 'субботу'][self.day]}\n"
                                if self.day == 0:
                                    schedule.week += 1
                                if schedule.week % 2 == 1:
                                    j = 1
                                    for i in range(0, 12, 2):
                                        message += f"{j}) {sch[i + self.day * 12]}\n"
                                        j += 1
                                else:
                                    j = 1
                                    for i in range(1, 12, 2):
                                        message += f"{j}) {sch[i + self.day * 12]}\n"
                                        j += 1
                                vk.messages.send(
                                    user_id=event.user_id,
                                    random_id=get_random_id(),
                                    message=message
                                )
                                break
                            elif event_c.text == "на эту неделю":
                                message += f"расписание {schedule.teacher} на {schedule.week} неделю\n"
                                if schedule.week % 2 == 1:
                                    for d in range(6):
                                        j = 1
                                        message += ["понедельник", 'вторник', "среда", "четверг", "пятница", "суббота"][
                                                       d] + "\n"
                                        for i in range(0, 12, 2):
                                            message += f"{j}) {sch[i + d * 12]}\n"
                                            j += 1
                                else:
                                    for d in range(6):
                                        j = 1
                                        message += ["понедельник", 'вторник', "среда", "четверг", "пятница", "суббота"][
                                                       d] + "\n"
                                        for i in range(1, 12, 2):
                                            message += f"{j}) {sch[i + d * 12]}\n"
                                            j += 1
                                vk.messages.send(
                                    user_id=event.user_id,
                                    random_id=get_random_id(),
                                    message=message
                                )
                                break
                            elif event_c.text == "на следующую неделю":
                                schedule.week += 1
                                message += f"расписание {schedule.teacher} на {schedule.week} неделю\n"
                                if schedule.week % 2 == 1:
                                    for d in range(6):
                                        j = 1
                                        message += ["понедельник", 'вторник', "среда", "четверг", "пятница", "суббота"][
                                                       d] + "\n"
                                        for i in range(0, 12, 2):
                                            message += f"{j}) {sch[i + d * 12]}\n"
                                            j += 1
                                else:
                                    for d in range(6):
                                        j = 1
                                        message += ["понедельник", 'вторник', "среда", "четверг", "пятница", "суббота"][
                                                       d] + "\n"
                                        for i in range(1, 12, 2):
                                            message += f"{j}) {sch[i + d * 12]}\n"
                                            j += 1
                                vk.messages.send(
                                    user_id=event.user_id,
                                    random_id=get_random_id(),
                                    message=message
                                )
                                break
                            else:
                                vk.messages.send(
                                    user_id=event.user_id,
                                    random_id=get_random_id(),
                                    message='некорректный запрос'
                                )
                                break

            else:
                vk.messages.send(
                    user_id=event.user_id,
                    random_id=get_random_id(),
                    message='такого преподавателя не существует'
                )
            self.logging("SENT", "teacher's schedule message")

        # get covid for russia
        elif message.upper() == "КОРОНА":
            corona = Corona("")
            corona.get_russia_covid()

            message = 'По состоянию на ' + corona.date + \
                      '\nСлучаев: ' + corona.cases + ' (' + corona.new_cases + ' за сегодня)' + \
                      '\nАктивных: ' + corona.active + ' (' + corona.new_active + ' за сегодня)' + \
                      '\nВылечено: ' + corona.cured + ' (' + corona.new_cured + ' за сегодня)' + \
                      '\nУмерло: ' + corona.died + ' (' + corona.new_died + ' за сегодня)'

            corona.get_covid_stat()
            upload = VkUpload(vk_session)
            attachments = []
            photo = upload.photo_messages(photos="corona.jpg")[0]
            attachments.append("photo{}_{}".format(photo["owner_id"], photo["id"]))

            vk.messages.send(
                user_id=event.user_id,
                attachment=','.join(attachments),
                random_id=get_random_id(),
                message=message
            )
            self.logging("SENT", "corona stat message")
        # get covid for any region
        elif re.fullmatch("КОРОНА " + r"[А-Я]+", message.upper()):
            region = message[8:].lower()
            corona = Corona(region)
            if corona.get_corona_region() == -1:
                message = "регион не найден"
            else:
                message = 'По состоянию на ' + corona.date + \
                          '\nрегион: ' + corona.region + \
                          '\nСлучаев: ' + corona.cases + ' (' + corona.new_cases + ' за сегодня)' + \
                          '\nАктивных: ' + corona.active + ' (' + corona.new_active + ' за сегодня)' + \
                          '\nВылечено: ' + corona.cured + ' (' + corona.new_cured + ' за сегодня)' + \
                          '\nУмерло: ' + corona.died + ' (' + corona.new_died + ' за сегодня)'

            vk.messages.send(
                user_id=event.user_id,
                random_id=get_random_id(),
                message=message
            )
            self.logging("SENT", "corona stat message")


        # no such command
        else:
            vk.messages.send(
                user_id=event.user_id,
                random_id=get_random_id(),
                message="некорректный запрос"
            )

    # for cleaning string from tags
    @staticmethod
    def _clean_all_tag_from_str(string_line):
        result = ""
        not_skip = True
        for i in list(string_line):
            if not_skip:
                if i == "<":
                    not_skip = False
                else:
                    result += i
            else:
                if i == ">":
                    not_skip = True

        return result

    # setting keyboard (schedule choice) key = 0 - teacher, 1 - student
    @staticmethod
    def set_keyboard_schedule(key, group="", teacher=""):
        keyboard = VkKeyboard(one_time=True)
        keyboard.add_button('на сегодня', color=VkKeyboardColor.POSITIVE)
        keyboard.add_button('на завтра', color=VkKeyboardColor.NEGATIVE)
        keyboard.add_line()
        keyboard.add_button('на эту неделю', color=VkKeyboardColor.PRIMARY)
        keyboard.add_button('на следующую неделю', color=VkKeyboardColor.PRIMARY)
        if key:
            keyboard.add_line()
            keyboard.add_button('какая неделя?', color=VkKeyboardColor.SECONDARY)
            keyboard.add_button('какая группа?', color=VkKeyboardColor.SECONDARY)
            vk.messages.send(
                keyboard=keyboard.get_keyboard(),
                user_id=event.user_id,
                random_id=get_random_id(),
                message=f'показать расписание группы {group} ...'
            )
            return
        vk.messages.send(
            keyboard=keyboard.get_keyboard(),
            user_id=event.user_id,
            random_id=get_random_id(),
            message=f'показать расписание преподавателя {teacher} ...'
        )

    # setting keyboard (weather time choice)
    @staticmethod
    def set_keyboard_weather():
        keyboard = VkKeyboard(one_time=True)
        keyboard.add_button('сейчас', color=VkKeyboardColor.PRIMARY)
        keyboard.add_button('на сегодня', color=VkKeyboardColor.POSITIVE)
        keyboard.add_button('на завтра', color=VkKeyboardColor.POSITIVE)
        keyboard.add_line()
        keyboard.add_button('на 5 дней', color=VkKeyboardColor.POSITIVE)
        vk.messages.send(
            keyboard=keyboard.get_keyboard(),
            user_id=event.user_id,
            random_id=get_random_id(),
            message='показать погоду в Москве'
        )

    # setting keyboard (teacher's surnames match)
    @staticmethod
    def set_keyboard_surnames(teachers):
        keyboard = VkKeyboard(one_time=True)
        for teacher in teachers:
            keyboard.add_button(f"Найти {teacher}", color=VkKeyboardColor.PRIMARY)
            if teacher != teachers[-1]:
                keyboard.add_line()
        vk.messages.send(
            keyboard=keyboard.get_keyboard(),
            user_id=event.user_id,
            random_id=get_random_id(),
            message='выберите преподавателя'
        )


# creating .xlsx files
def get_schedule_files():
    schedule_request = requests.get("https://www.mirea.ru/schedule/")
    schedule_soup = bs.BeautifulSoup(schedule_request.text, "html.parser")
    schedule_result = schedule_soup.find_all('div', id="toggle-hl_2_1-hl_3_3")

    links = re.findall(r"https.+21-22_весна_очка.xlsx", str(schedule_result))
    with open('links.txt', 'r') as f:
        links_old = f.readlines()

    # updating links if needed
    if len(links_old) < 3 or links_old[0] != links[0] + '\n' or links_old[1] != links[1] + '\n' and links_old[2] != \
            links[2] + '\n':
        links_f = open("links.txt", "w")
        for list_n in range(len(links)):
            f = open(f"course{list_n + 1}.xlsx", "wb")
            resp = requests.get(links[list_n])
            links_f.write(links[list_n] + '\n')
            f.write(resp.content)
            f.close()
        links_f.close()

        # opening all courses
        book1 = openpyxl.load_workbook("course1.xlsx")
        book2 = openpyxl.load_workbook("course2.xlsx")
        book3 = openpyxl.load_workbook("course3.xlsx")
        global course1, course2, course3
        course1 = book1.active
        course2 = book2.active
        course3 = book3.active


f = open("links.txt", "a")
f.close()
f = open("log.txt", 'w')
f.close()
f = shelve.open("groups.txt", 'c')
f.close()

get_schedule_files()
book1 = openpyxl.load_workbook("course1.xlsx")
book2 = openpyxl.load_workbook("course2.xlsx")
book3 = openpyxl.load_workbook("course3.xlsx")
course1 = book1.active
course2 = book2.active
course3 = book3.active

# Авторизуемся как сообщество
vk_session = vk_api.VkApi(token=VK_API_TOKEN)

vk = vk_session.get_api()
# Работа с сообщениями
longpoll = VkLongPoll(vk_session)

print("Server started")
messages = dict()
for event in longpoll.listen():
    if event.type == VkEventType.MESSAGE_NEW:

        if event.to_me:
            bot = VkBot(event.user_id)

            messages[event.user_id] = event.text

            bot.logging("NEW", f'new message from {event.user_id}')
            print(f'New message from {event.user_id}')

            bot.new_message(messages[event.user_id])

            print('Text: ', event.text)
            print("-------------------")
