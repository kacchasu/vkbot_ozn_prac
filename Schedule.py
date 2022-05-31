import requests
import bs4 as bs
import re
from datetime import datetime, time, date
import openpyxl


class Schedule:
    def __init__(self, group_id, teacher_id, course_1, course_2, course_3):
        self.course1 = course_1
        self.course2 = course_2
        self.course3 = course_3

        self.week = self.get_week_num()

        self.group = group_id
        self.teacher = teacher_id

        self.teacher_day = [""] * 6
        self.student_day = [""] * 6

        self.logging("CRE", "schedule class object: created")

    # for logging
    @staticmethod
    def logging(key, comm):
        f1 = open("log.txt", "a")
        t1 = datetime.now()
        f1.write(key + " --- ")
        f1.write(t1.strftime("%d/%m/%Y %H:%M") + " --- ")
        f1.write("schedule class: " + comm + '\n')

    # for getting teacher surname matches
    def check_surnames(self):
        num_cols1 = self.course1.max_column
        num_cols2 = self.course2.max_column
        num_cols3 = self.course3.max_column
        num_rows1 = self.course1.max_row
        num_rows2 = self.course2.max_row
        num_rows3 = self.course3.max_row

        teachers = set()

        # checking course 1
        for column in range(8, num_cols1, 5):
            for row in range(4, num_rows1):
                temp = str(self.course1.cell(row=row, column=column).value)
                if temp.find(self.teacher) > -1 and temp[temp.find(self.teacher) + len(self.teacher)] == " ":
                    teachers.add(self.course1.cell(row=row, column=column).value[
                                 temp.find(self.teacher):temp.find(self.teacher) + len(self.teacher) + 5])

        # checking course 2
        for column in range(8, num_cols2, 5):
            for row in range(4, num_rows2):
                temp = str(self.course2.cell(row=row, column=column).value)
                if temp.find(self.teacher) > -1 and temp[temp.find(self.teacher) + len(self.teacher)] == " ":
                    teachers.add(self.course2.cell(row=row, column=column).value[
                                 temp.find(self.teacher):temp.find(self.teacher) + len(self.teacher) + 5])

        # checking course 3
        for column in range(8, num_cols3, 5):
            for row in range(4, num_rows3):
                temp = str(self.course3.cell(row=row, column=column).value)
                if temp.find(self.teacher) > -1 and temp[temp.find(self.teacher) + len(self.teacher)] == " ":
                    teachers.add(self.course3.cell(row=row, column=column).value[
                                 temp.find(self.teacher):temp.find(self.teacher) + len(self.teacher) + 5])

        self.logging("CHECK", "schedule class object: surnames checked")

        return teachers

    # getting schedule week (current \ next)
    def get_week_schedule(self):
        num_cols1 = self.course1.max_column
        num_cols2 = self.course2.max_column
        num_cols3 = self.course3.max_column

        week_schedule = ["--"] * 72
        day = datetime.now().weekday()

        # if teacher in context
        if self.teacher != "":

            # checking course 1
            for column in range(6, num_cols1, 5):
                for row in range(4, 76):
                    temp = str(self.course1.cell(row=row, column=column + 2).value)
                    if temp.find(self.teacher) > -1:
                        cell2 = str(self.course1.cell(row=row, column=column + 1).value) if str(
                            self.course1.cell(row=row,
                                              column=column + 1).value) != "" else "--"
                        cell3 = str(self.course1.cell(row=row, column=column + 2).value) if str(
                            self.course1.cell(row=row,
                                              column=column + 2).value) != "" else "--"
                        cell4 = str(self.course1.cell(row=row, column=column + 3).value) if str(
                            self.course1.cell(row=row,
                                              column=column + 3).value) != "" else "--"
                        week_schedule[row - 4] = f'{str(self.course1.cell(row=row, column=column).value)}, ' \
                                                 f'{cell2}, {cell3}, {cell4}'

            # checking course 2
            for column in range(6, num_cols2, 5):
                for row in range(4, 76):
                    temp = str(self.course2.cell(row=row, column=column + 2).value)
                    if temp.find(self.teacher) > -1:
                        cell2 = str(self.course2.cell(row=row, column=column + 1).value) if str(
                            self.course2.cell(row=row,
                                              column=column + 1).value) != "" else "--"
                        cell3 = str(self.course2.cell(row=row, column=column + 2).value) if str(
                            self.course2.cell(row=row,
                                              column=column + 2).value) != "" else "--"
                        cell4 = str(self.course2.cell(row=row, column=column + 3).value) if str(
                            self.course2.cell(row=row,
                                              column=column + 3).value) != "" else "--"
                        week_schedule[row - 4] = f'{str(self.course2.cell(row=row, column=column).value)}, ' \
                                                 f'{cell2}, {cell3}, {cell4}'

            # checking course 3
            for column in range(6, num_cols3, 5):
                for row in range(4, 76):
                    temp = str(self.course3.cell(row=row, column=column + 2).value)
                    if temp.find(self.teacher) > -1:
                        cell2 = str(self.course3.cell(row=row, column=column + 1).value) if str(
                            self.course3.cell(row=row,
                                              column=column + 1).value) != "" else "--"
                        cell3 = str(self.course3.cell(row=row, column=column + 2).value) if str(
                            self.course3.cell(row=row,
                                              column=column + 2).value) != "" else "--"
                        cell4 = str(self.course3.cell(row=row, column=column + 3).value) if str(
                            self.course3.cell(row=row,
                                              column=column + 3).value) != "" else "--"
                        week_schedule[row - 4] = f'{str(self.course3.cell(row=row, column=column).value)}, ' \
                                                 f'{cell2}, {cell3}, {cell4}'

        # if group in context
        if self.group != "":

            # checking course 1
            if self.group[len(self.group) - 2:len(self.group)] == "21":
                for column in range(6, num_cols1, 5):
                    if str(self.course1.cell(row=2, column=column).value) == self.group:
                        for row in range(4, 76):
                            if self.course1.cell(row=row, column=column).value is not None:
                                cell2 = str(self.course1.cell(row=row, column=column + 1).value) if self.course1.cell(
                                    row=row,
                                    column=column + 1).value is not None else "--"
                                cell3 = str(self.course1.cell(row=row, column=column + 2).value) if self.course1.cell(
                                    row=row,
                                    column=column + 2).value is not None else "--"
                                cell4 = str(self.course1.cell(row=row, column=column + 3).value) if self.course1.cell(
                                    row=row,
                                    column=column + 3).value is not None else "--"
                                week_schedule[row - 4] = f'{self.course1.cell(row=row, column=column).value}, ' \
                                                         f'{cell2}, {cell3}, {cell4}'
                        break

            # checking course 2
            elif self.group[len(self.group) - 2:len(self.group)] == "20":
                for column in range(6, num_cols2, 5):
                    if str(self.course2.cell(row=2, column=column).value) == self.group:
                        for row in range(4, 76):
                            if self.course2.cell(row=row, column=column).value is not None:
                                cell2 = str(self.course2.cell(row=row, column=column + 1).value) if self.course2.cell(
                                    row=row,
                                    column=column + 1).value is not None else "--"
                                cell3 = str(self.course2.cell(row=row, column=column + 2).value) if self.course2.cell(
                                    row=row,
                                    column=column + 2).value is not None else "--"
                                cell4 = str(self.course2.cell(row=row, column=column + 3).value) if self.course2.cell(
                                    row=row,
                                    column=column + 3).value is not None else "--"
                                week_schedule[row - 4] = f'{self.course2.cell(row=row, column=column).value}, ' \
                                                         f'{cell2}, {cell3}, {cell4}'
                        break

            # checking course 3
            elif self.group[len(self.group) - 2:len(self.group)] == "19":
                for column in range(6, num_cols3, 5):
                    if str(self.course3.cell(row=2, column=column).value) == self.group:
                        for row in range(4, 76):
                            if self.course3.cell(row=row, column=column).value is not None:
                                cell2 = str(self.course3.cell(row=row, column=column + 1).value) if self.course3.cell(
                                    row=row,
                                    column=column + 1).value is not None else "--"
                                cell3 = str(self.course3.cell(row=row, column=column + 2).value) if self.course3.cell(
                                    row=row,
                                    column=column + 2).value is not None else "--"
                                cell4 = str(self.course3.cell(row=row, column=column + 3).value) if self.course3.cell(
                                    row=row,
                                    column=column + 3).value is not None else "--"
                                week_schedule[row - 4] = f'{self.course3.cell(row=row, column=column).value}, ' \
                                                         f'{cell2}, {cell3}, {cell4}'
                        break

        self.logging("GET", "schedule class object: got week schedule")

        return week_schedule

    # getting week number
    def get_week_num(self):
        sem_start = datetime(2022, 2, 9)
        today = datetime.now()
        days_ = today - sem_start
        week_num = days_.days // 7 + 1
        if today.weekday() < 2:
            week_num += 1
        self.logging("GET", "schedule class object: got week number")

        return week_num
